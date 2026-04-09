#!/usr/bin/env python3
"""
Drake Tax Return Pre-Processor

Extracts data from prior year tax return PDFs and current year Excel/CSV files,
then outputs a structured JSON file mapped to Drake screen fields.

Usage:
    python preprocess.py \
        --prior-year-pdf <path> \
        --current-year-data <path> \
        --return-type 1065 \
        --output return_data.json

Supported source formats:
    - Prior year: PDF (Drake-generated tax return)
    - Current year: .xlsx, .xls, .csv

The script uses pattern matching to extract data from PDF text. It works best
with Drake-generated prior year returns (consistent formatting). For non-Drake
prior year returns, some fields may need manual entry in the output JSON.
"""

import argparse
import json
import re
import sys
import os
from datetime import date

# ---------------------------------------------------------------------------
# PDF text extraction
# ---------------------------------------------------------------------------

def extract_pdf_text(pdf_path: str) -> str:
    """Extract all text from a PDF. Tries pdfplumber first, falls back to PyPDF2."""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)
    except ImportError:
        pass

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(pdf_path)
        text_parts = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text_parts.append(t)
        return "\n".join(text_parts)
    except ImportError:
        pass

    print("ERROR: No PDF library available. Install pdfplumber or PyPDF2:")
    print("  pip install pdfplumber --break-system-packages")
    sys.exit(1)


def extract_pdf_tables(pdf_path: str) -> list:
    """Extract tables from PDF using pdfplumber."""
    try:
        import pdfplumber
        tables = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)
        return tables
    except ImportError:
        return []


# ---------------------------------------------------------------------------
# Excel / CSV reading
# ---------------------------------------------------------------------------

def read_excel(path: str) -> dict:
    """Read an Excel file and return a dict of sheet_name -> list of row dicts."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                row_dict = {}
                for i, val in enumerate(row):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    row_dict[key] = val
                data.append(row_dict)
            result[sheet_name] = data
        return result
    except ImportError:
        print("ERROR: openpyxl not available. Install it:")
        print("  pip install openpyxl --break-system-packages")
        sys.exit(1)


def read_csv(path: str) -> list:
    """Read a CSV file and return a list of row dicts."""
    import csv
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return list(reader)


# ---------------------------------------------------------------------------
# 1065 Prior Year PDF Extraction
# ---------------------------------------------------------------------------

def extract_1065_from_pdf(text: str, tables: list) -> dict:
    """
    Extract 1065 partnership data from a Drake-generated prior year return PDF.

    Drake PDFs have a specific structure:
    - Diagnostic Report page (page 1)
    - Cover letter (page ~11)
    - Form 8825 (Rental Real Estate)
    - Form 4562 (Depreciation)
    - Schedule K-1s (one per partner, preceded by a partner cover letter)
    - Comparison Schedule (prior year vs current year)

    The extraction uses patterns specific to Drake's PDF formatting. Non-Drake
    PDFs may require different patterns or manual JSON construction.
    """
    data = {
        "return_type": "1065",
        "tax_year": None,
        "entity": {},
        "rental_properties": [],
        "depreciation_assets": [],
        "partners": [],
        "schedule_b": {"all_no": True},
        "sec": {},
        "schedule_m1": {},
        "k_credits": {"k2_k3_not_required": True},
        "pin": {
            "ero_pin": "75757",
            "ero_entered": True,
        },
    }

    # --- Tax year (from Diagnostic Report or cover letter) ---
    year_match = re.search(r"Tax Year\s*:\s*(\d{4})", text)
    if year_match:
        data["tax_year"] = int(year_match.group(1)) + 1  # next year's return
    else:
        year_match = re.search(r"(?:calendar year|tax year|year ended[^\d]*)(\d{4})", text, re.IGNORECASE)
        if year_match:
            data["tax_year"] = int(year_match.group(1)) + 1

    # --- Entity info from Diagnostic Report header ---
    # Drake Diagnostic Report format: "Taxpayer: NAME\nID No : XX-XXXXXXX"
    taxpayer_match = re.search(r"Taxpayer:\s*(.+?)(?:\n|$)", text)
    if taxpayer_match:
        data["entity"]["name"] = taxpayer_match.group(1).strip()

    ein_match = re.search(r"ID No\s*:\s*(\d{2}-\d{7})", text)
    if not ein_match:
        ein_match = re.search(r"(\d{2}-\d{7})", text)
    if ein_match:
        data["entity"]["ein"] = ein_match.group(1)

    # --- Address from cover letter ---
    # Drake cover letter format: "ENTITY NAME\nSTREET\nCITY, ST ZIP"
    entity_name = data["entity"].get("name", "")
    if entity_name:
        addr_pattern = re.compile(
            re.escape(entity_name) + r"\s*\n"
            r"(.+?)\n"
            r"([A-Z][A-Z\s]+?),\s*([A-Z]{2})\s+(\d{5})",
            re.MULTILINE
        )
        addr_match = addr_pattern.search(text)
        if addr_match:
            data["entity"]["address"] = {
                "street": addr_match.group(1).strip(),
                "city": addr_match.group(2).strip(),
                "state": addr_match.group(3),
                "zip": addr_match.group(4),
            }

    # --- Partner info from K-1 cover letters ---
    # Drake generates a cover letter before each K-1 with format:
    # "Dear PARTNER NAME:\n...your XX.XXXXXX percent interest in:\n...\n
    #  Capital Account at Beginning of Year $AMOUNT"
    partner_letters = re.finditer(
        r"Dear\s+([A-Z][A-Z\s]+?):\s*\n"
        r".*?(\d+\.\d+)\s*percent interest"
        r".*?Capital Account at Beginning of Year\s*\$?([-\d,]+)",
        text, re.DOTALL
    )
    seen_partners = set()
    for m in partner_letters:
        name = m.group(1).strip()
        if name in seen_partners:
            continue
        seen_partners.add(name)

        pct = float(m.group(2))
        beg_capital = int(m.group(3).replace(",", ""))

        # Try to find SSN and address from the same K-1 block
        # Look for SSN pattern near this partner's name
        block_start = m.start()
        block_text = text[max(0, block_start - 200):block_start + 2000]

        ssn_match = re.search(r"(\d{3}-\d{2}-\d{4})", block_text)
        ssn = ssn_match.group(1) if ssn_match else ""

        # Address: look for lines before "Dear NAME:"
        addr_lines = re.search(
            r"([A-Z][A-Z\s]+?)\n"
            r"(\d+.+?)\n"
            r"([A-Z][A-Z\s]+?),\s*([A-Z]{2})\s+(\d{5})\s*\n"
            r".*?Dear\s+" + re.escape(name),
            text[max(0, block_start - 500):block_start + 500],
            re.DOTALL
        )

        partner = {
            "name": name,
            "ssn": ssn,
            "ownership_pct": round(pct),
            "profit_pct": round(pct),
            "loss_pct": round(pct),
            "capital_pct": round(pct),
            "beginning_capital": beg_capital,
            "partner_type": "general",
            "is_50_pct_or_more": pct >= 50,
            "country": "US",
        }

        if addr_lines:
            partner["address"] = {
                "street": addr_lines.group(2).strip(),
                "city": addr_lines.group(3).strip(),
                "state": addr_lines.group(4),
                "zip": addr_lines.group(5),
            }

        data["partners"].append(partner)

    # --- 8825 Rental Real Estate data ---
    # Look for Form 8825 section
    f8825_match = re.search(r"8825.*?Rental Real Estate", text)
    if f8825_match:
        # Find the property address on 8825
        prop_addr_match = re.search(
            r"8825.*?(\d+\s+[A-Z][\w\s]+(?:Ave|St|Rd|Dr|Blvd|Ln|Way|Ct|Pl|Cir)[^\n]*)\n\s*([A-Z][\w\s]+?),\s*([A-Z]{2})\s+(\d{5})",
            text, re.DOTALL
        )
        prop = {"expenses": {}}
        if prop_addr_match:
            prop["address"] = {
                "street": prop_addr_match.group(1).strip(),
                "city": prop_addr_match.group(2).strip(),
                "state": prop_addr_match.group(3),
                "zip": prop_addr_match.group(4),
            }

        # Extract 8825 line amounts using line number patterns
        line_patterns = {
            "gross_rents": r"(?:Gross rents|line 2[^0-9])\s*2?\s*([\d,]+)\.",
            "commissions": r"(?:Commission|line 6[^0-9])\s*6?\s*([\d,]+)\.",
            "insurance": r"(?:Insurance|line 7[^0-9])\s*7?\s*([\d,]+)\.",
            "legal_professional": r"(?:Legal|line 8[^0-9])\s*8?\s*([\d,]+)\.",
            "interest_mortgage": r"(?:Interest.*?instruction|line 9[^0-9])\s*9?\s*([\d,]+)\.",
            "repairs": r"(?:Repair|line 10[^0-9])\s*10?\s*([\d,]+)\.",
            "taxes": r"(?:Taxes|line 11[^0-9])\s*11?\s*([\d,]+)\.",
            "utilities": r"(?:Utilit|line 12[^0-9])\s*12?\s*([\d,]+)\.",
            "wages": r"(?:Wages|line 13[^0-9])\s*13?\s*([\d,]+)\.",
        }

        # Search within the 8825 section (roughly 3000 chars after the header)
        section_text = text[f8825_match.start():f8825_match.start() + 4000]

        for field, pattern in line_patterns.items():
            m = re.search(pattern, section_text, re.IGNORECASE)
            if m:
                val = int(m.group(1).replace(",", ""))
                if field == "gross_rents":
                    prop["gross_rents"] = val
                else:
                    prop["expenses"][field] = val

        prop["qbi_trade_or_business"] = "Y"
        data["rental_properties"].append(prop)

    # --- Depreciation from 4562 and Diagnostic Report ---
    # Drake Diagnostic Report often lists assets clearly:
    # "Description: Building Date Placed in Service: 08/25/2017\nCost: 1,097,209."
    diag_assets = re.finditer(
        r"Description:\s*(\w[\w\s]*?)\s*Date Placed in Service:\s*(\d{2}/\d{2}/\d{4})\s*\n\s*Cost:\s*([\d,]+)",
        text
    )
    seen_assets = set()
    for m in diag_assets:
        desc = m.group(1).strip()
        if desc in seen_assets:
            continue
        seen_assets.add(desc)
        asset = {
            "description": desc,
            "date_in_service": m.group(2),
            "cost_basis": int(m.group(3).replace(",", "")),
            "method": "ARR",  # Default to MACRS residential rental
            "life": 27.5,
            "for_form": "8825",
        }
        data["depreciation_assets"].append(asset)

    # Also try the 4562 depreciation statement format (if present)
    # Drake 4562 statements often appear as tables
    if not data["depreciation_assets"]:
        # Fallback: look for any "Building" + date + cost pattern
        asset_pattern = re.compile(
            r"(Building|Unit Remodel|Improvements|Equipment|Furniture|Appliances)"
            r".*?(\d{2}/\d{2}/\d{4}).*?([\d,]+)\.",
            re.DOTALL
        )
        for m in asset_pattern.finditer(text):
            desc = m.group(1).strip()
            if desc in seen_assets:
                continue
            seen_assets.add(desc)
            asset = {
                "description": desc,
                "date_in_service": m.group(2),
                "cost_basis": int(m.group(3).replace(",", "")),
                "method": "ARR",
                "life": 27.5,
                "for_form": "8825",
            }
            data["depreciation_assets"].append(asset)

    # --- Prior year income for SEC screen (from Comparison Schedule) ---
    # Drake PDFs often have OCR artifacts (random 'm' chars) in text.
    # The Comparison Schedule line looks like:
    # "Net income (loss) from rental realm emstmatme macmtivmitimesm  -32,208. -93,844. 61,636."
    # We use a loose pattern that tolerates these artifacts.
    sec_match = re.search(
        r"Net income \(loss\) from rental re.*?\s+([-]?[\d,]+)\.\s",
        text
    )
    if sec_match:
        val_str = sec_match.group(1).replace(",", "")
        data["sec"]["prior_year_income"] = int(val_str)

    # --- M-1 basis ---
    if re.search(r"Tax.?[Bb]asis", text):
        data["schedule_m1"]["basis"] = "tax"
    else:
        data["schedule_m1"]["basis"] = "tax"  # default for rental partnerships

    # --- State info ---
    # Check if IL return is present
    if re.search(r"Illinois", text, re.IGNORECASE):
        data["entity"]["resident_state"] = "IL"
        # Pre-populate IL state requirements
        entity_addr = data["entity"].get("address", {})
        if entity_addr:
            data["state_il"] = {
                "books_city": entity_addr.get("city", ""),
                "books_state": entity_addr.get("state", ""),
                "books_zip": entity_addr.get("zip", ""),
                "principal_place_city": entity_addr.get("city", ""),
                "principal_place_state": entity_addr.get("state", ""),
                "principal_place_zip": entity_addr.get("zip", ""),
                "il_authorization": True,
                "create_ilk1": True,
            }

    return data


# ---------------------------------------------------------------------------
# Current Year Excel Processing (for rental income/expenses)
# ---------------------------------------------------------------------------

# Common mappings from Excel column names to Drake expense fields
EXPENSE_MAPPING = {
    # Standard 8825 expense lines
    "commissions": "commissions",
    "commission": "commissions",
    "insurance": "insurance",
    "interest": "interest_mortgage",
    "mortgage interest": "interest_mortgage",
    "interest - mortgage": "interest_mortgage",
    "legal": "legal_professional",
    "legal & professional": "legal_professional",
    "legal and professional": "legal_professional",
    "management": "management_fees",
    "management fees": "management_fees",
    "property management": "management_fees",
    "repairs": "repairs",
    "repairs & maintenance": "repairs",
    "repairs and maintenance": "repairs",
    "taxes": "taxes",
    "real estate taxes": "taxes",
    "property taxes": "taxes",
    "re taxes": "taxes",
    "utilities": "utilities",
    "wages": "wages",
    "salaries": "wages",
    "payroll": "wages",
    "advertising": "advertising",
    "auto": "auto_travel",
    "travel": "auto_travel",
    "cleaning": "cleaning_maintenance",
    "supplies": "supplies",
}


def normalize_key(s: str) -> str:
    """Normalize a string for matching against expense mappings."""
    return re.sub(r"[^a-z0-9 ]", "", s.lower()).strip()


def process_current_year_excel(excel_data: dict) -> dict:
    """
    Process current year Excel data into rental property income/expenses.

    Handles multiple formats:
    - Single sheet with income/expense rows
    - Multiple sheets (one per property)
    - Column-based comparison format (property in columns)
    """
    result = {
        "rental_properties": [],
        "gross_rents": None,
        "expenses": {},
        "other_expenses": [],
    }

    # Try each sheet
    for sheet_name, rows in excel_data.items():
        if not rows:
            continue

        # Detect format: look at headers
        headers = list(rows[0].keys()) if rows else []

        # Format 1: Row-based (Description | Amount columns)
        if any("description" in normalize_key(h) for h in headers) or \
           any("category" in normalize_key(h) for h in headers):
            result = _process_row_based(rows, headers)

        # Format 2: Column-based comparison (Account | Property1 | Property2 | Total)
        elif any("total" in normalize_key(h) for h in headers) or \
             any("account" in normalize_key(h) for h in headers) or len(headers) >= 3:
            result = _process_column_based(rows, headers)

        # Format 3: Simple two-column (Label | Value)
        elif len(headers) == 2:
            result = _process_two_column(rows, headers)

    return result


def _process_column_based(rows: list, headers: list) -> dict:
    """Process column-based format where each property is a column."""
    result = {"expenses": {}, "other_expenses": [], "gross_rents": None}

    # Find the label column (first text column) and value columns
    label_col = headers[0]

    # Use Total column if it exists, otherwise last numeric column
    value_col = None
    for h in headers:
        if "total" in normalize_key(h):
            value_col = h
            break
    if not value_col:
        value_col = headers[-1]  # fallback to last column

    for row in rows:
        label = str(row.get(label_col, "")).strip()
        if not label:
            continue

        raw_val = row.get(value_col)
        if raw_val is None:
            continue

        # Parse numeric value
        try:
            if isinstance(raw_val, (int, float)):
                val = round(abs(raw_val))
            else:
                cleaned = str(raw_val).replace(",", "").replace("$", "").replace("(", "-").replace(")", "").strip()
                if not cleaned or cleaned == "-":
                    continue
                val = round(abs(float(cleaned)))
        except (ValueError, TypeError):
            continue

        if val == 0:
            continue

        normalized = normalize_key(label)

        # Check for income/rent
        if any(kw in normalized for kw in ["rent", "income", "revenue", "gross"]):
            if result["gross_rents"] is None or val > result["gross_rents"]:
                result["gross_rents"] = val
            continue

        # Map to standard expense field
        mapped = EXPENSE_MAPPING.get(normalized)
        if not mapped:
            # Try partial matching
            for key, field in EXPENSE_MAPPING.items():
                if key in normalized or normalized in key:
                    mapped = field
                    break

        if mapped:
            result["expenses"][mapped] = result["expenses"].get(mapped, 0) + val
        else:
            # Treat as "other expense"
            result["other_expenses"].append({
                "description": label,
                "amount": val,
            })

    return result


def _process_row_based(rows: list, headers: list) -> dict:
    """Process row-based format with description and amount columns."""
    result = {"expenses": {}, "other_expenses": [], "gross_rents": None}

    desc_col = None
    amt_col = None
    for h in headers:
        nh = normalize_key(h)
        if "desc" in nh or "category" in nh or "account" in nh or "item" in nh:
            desc_col = h
        elif "amount" in nh or "total" in nh or "value" in nh:
            amt_col = h

    if not desc_col:
        desc_col = headers[0]
    if not amt_col:
        amt_col = headers[-1]

    for row in rows:
        label = str(row.get(desc_col, "")).strip()
        raw_val = row.get(amt_col)
        if not label or raw_val is None:
            continue

        try:
            if isinstance(raw_val, (int, float)):
                val = round(abs(raw_val))
            else:
                cleaned = str(raw_val).replace(",", "").replace("$", "").strip()
                val = round(abs(float(cleaned)))
        except (ValueError, TypeError):
            continue

        if val == 0:
            continue

        normalized = normalize_key(label)

        if any(kw in normalized for kw in ["rent", "income", "revenue"]):
            result["gross_rents"] = val
            continue

        mapped = EXPENSE_MAPPING.get(normalized)
        if mapped:
            result["expenses"][mapped] = val
        else:
            result["other_expenses"].append({"description": label, "amount": val})

    return result


def _process_two_column(rows: list, headers: list) -> dict:
    """Process simple two-column label/value format."""
    return _process_row_based(rows, headers)


# ---------------------------------------------------------------------------
# Merge prior year + current year into final JSON
# ---------------------------------------------------------------------------

def clean_other_expenses(expenses: list) -> list:
    """Remove subtotal/total rows from other expenses list."""
    skip_keywords = ["total", "net income", "net loss", "gross", "noi"]
    cleaned = []
    for exp in expenses:
        desc_lower = exp.get("description", "").lower()
        if any(kw in desc_lower for kw in skip_keywords):
            continue
        cleaned.append(exp)
    return cleaned


def clean_address_field(value: str) -> str:
    """Remove PDF artifacts from address fields."""
    if not value:
        return value
    # Remove Drake PDF footer lines (e.g., "9892TY WK4Q 04/08/2026...")
    value = re.sub(r"\d{4}TY\s+\w+\s+\d{2}/\d{2}/\d{4}.*?\n?", "", value)
    # Remove stray line labels (A, B, C, D from 8825 property rows)
    value = re.sub(r"^[A-D]\n", "", value)
    # Remove duplicate names that got captured
    lines = [l.strip() for l in value.split("\n") if l.strip()]
    # If multiple lines, take the last relevant one (most specific)
    if len(lines) > 1:
        # Keep only the street address line (has numbers)
        street_lines = [l for l in lines if re.search(r"\d", l)]
        if street_lines:
            return street_lines[-1]
    return lines[-1] if lines else value


def merge_data(prior_year: dict, current_year: dict, return_type: str) -> dict:
    """Merge prior year entity/depreciation data with current year financials."""
    result = prior_year.copy()

    # Update tax year to current
    if result.get("tax_year"):
        pass  # already set by extraction
    else:
        result["tax_year"] = date.today().year

    # Merge current year rental data
    if current_year.get("gross_rents") is not None:
        if result.get("rental_properties"):
            prop = result["rental_properties"][0]
            prop["gross_rents"] = current_year["gross_rents"]
            prop["expenses"] = current_year.get("expenses", {})
            if current_year.get("other_expenses"):
                prop["expenses"]["other_expenses"] = current_year["other_expenses"]
            prop.setdefault("qbi_trade_or_business", "Y")
        else:
            result["rental_properties"] = [{
                "gross_rents": current_year["gross_rents"],
                "expenses": current_year.get("expenses", {}),
                "qbi_trade_or_business": "Y",
            }]
            if current_year.get("other_expenses"):
                result["rental_properties"][0]["expenses"]["other_expenses"] = current_year["other_expenses"]

    # Ensure PIN defaults for Lineal CPA
    result.setdefault("pin", {})
    result["pin"].setdefault("ero_pin", "75757")
    result["pin"].setdefault("ero_entered", True)

    # Ensure depreciation assets have for_form set
    for asset in result.get("depreciation_assets", []):
        asset.setdefault("for_form", "8825")

    # Clean up other_expenses (remove subtotals)
    for prop in result.get("rental_properties", []):
        if "other_expenses" in prop.get("expenses", {}):
            prop["expenses"]["other_expenses"] = clean_other_expenses(
                prop["expenses"]["other_expenses"]
            )
        # Clean address fields
        if "address" in prop:
            for field in ["street", "city"]:
                if field in prop["address"]:
                    prop["address"][field] = clean_address_field(prop["address"][field])

    # Clean partner addresses
    for partner in result.get("partners", []):
        if "address" in partner:
            for field in ["street", "city"]:
                if field in partner["address"]:
                    partner["address"][field] = clean_address_field(partner["address"][field])

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Extract tax return data from source documents into Drake-ready JSON"
    )
    parser.add_argument("--prior-year-pdf", required=True, help="Path to prior year tax return PDF")
    parser.add_argument("--current-year-data", required=False, help="Path to current year Excel/CSV")
    parser.add_argument("--return-type", required=True, choices=["1065", "1120S", "1120", "1040"],
                        help="Type of tax return")
    parser.add_argument("--output", required=True, help="Output JSON file path")
    args = parser.parse_args()

    print(f"Pre-processing {args.return_type} return...")

    # --- Extract prior year PDF ---
    print(f"  Reading prior year PDF: {args.prior_year_pdf}")
    pdf_text = extract_pdf_text(args.prior_year_pdf)
    pdf_tables = extract_pdf_tables(args.prior_year_pdf)
    print(f"  Extracted {len(pdf_text)} chars of text, {len(pdf_tables)} tables")

    if args.return_type == "1065":
        prior_data = extract_1065_from_pdf(pdf_text, pdf_tables)
    else:
        print(f"  WARNING: Extraction for {args.return_type} is not yet implemented.")
        print(f"  Creating skeleton JSON — you'll need to fill in the data manually.")
        prior_data = {
            "return_type": args.return_type,
            "tax_year": date.today().year,
            "entity": {},
            "pin": {"ero_pin": "75757", "ero_entered": True},
        }

    # --- Process current year data ---
    current_data = {}
    if args.current_year_data:
        ext = os.path.splitext(args.current_year_data)[1].lower()
        print(f"  Reading current year data: {args.current_year_data}")

        if ext in (".xlsx", ".xls"):
            excel_data = read_excel(args.current_year_data)
            current_data = process_current_year_excel(excel_data)
        elif ext == ".csv":
            csv_data = {"Sheet1": read_csv(args.current_year_data)}
            current_data = process_current_year_excel(csv_data)
        else:
            print(f"  WARNING: Unsupported file format {ext}. Skipping current year data.")
    else:
        print("  No current year data file provided. Only prior year data will be used.")

    # --- Merge ---
    result = merge_data(prior_data, current_data, args.return_type)

    # --- Write output ---
    with open(args.output, "w") as f:
        json.dump(result, f, indent=2)

    print(f"\nOutput written to: {args.output}")
    print(f"Return type: {result.get('return_type')}")
    print(f"Tax year: {result.get('tax_year')}")
    print(f"Entity: {result.get('entity', {}).get('name', '(not extracted)')}")
    print(f"Partners: {len(result.get('partners', []))}")
    print(f"Depreciation assets: {len(result.get('depreciation_assets', []))}")
    print(f"Rental properties: {len(result.get('rental_properties', []))}")

    # --- Validation warnings ---
    warnings = []
    if not result.get("entity", {}).get("name"):
        warnings.append("Entity name not extracted — fill in manually")
    if not result.get("entity", {}).get("ein"):
        warnings.append("EIN not extracted — fill in manually")
    if not result.get("partners"):
        warnings.append("No partners extracted — fill in manually")
    if not result.get("depreciation_assets"):
        warnings.append("No depreciation assets extracted — check if applicable")
    for asset in result.get("depreciation_assets", []):
        if asset.get("for_form") != "8825":
            warnings.append(f"Asset '{asset.get('description')}' missing for_form='8825'")

    if warnings:
        print(f"\n⚠ {len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")
    else:
        print("\nNo warnings — data looks complete!")

    print("\nNext step: Review the JSON, fill in any missing fields, then use the")
    print("Drake Tax Return skill to enter data into Drake.")


if __name__ == "__main__":
    main()
